#!/usr/bin/python
# -*- coding: utf-8 -*-
import qrcode
import os
import requests
import arabic_reshaper
import persian
import time

from StringIO import StringIO
from openpyxl import load_workbook
from PIL import Image, ImageOps, ImageFont, ImageDraw

from bidi.algorithm import get_display

wb = load_workbook('attendees_list.xlsx')
ws = wb.get_sheet_by_name('report')
attendees = str(ws.max_row)
person_name = ws['F2':'F'+ attendees]

for row in ws.iter_rows('C{}:C{}'.format(ws.min_row+1,ws.max_row)): #ws.max_row
	for cell in row:
		# reset all values
		headshot = None
		qr = None
		qrimage = None
		text1 = None
		text2 = None
		reshaped_text = None
		final_text = None
		front = None
		back = None
		w1 = None
		w2 = None
		h1 = None
		h2 = None
		font1 = None
		font2 = None
		draw = None
		mask = None
		output = None
		# qr = qrcode.make()
		qr = qrcode.QRCode(
			version=1,
			error_correction=qrcode.constants.ERROR_CORRECT_L,
			box_size=12,
			border=4,
		)
		qr.add_data(cell.value)
		qr.make(fit=True)
		qrimage = qr.make_image(fill_color="black", back_color="white")
		back = Image.open("./PNG_badge/02 (2).png").convert("RGBA")
		back.paste(qrimage, (275, 360))
		back.save('./export/'+ws.cell(row=cell.row, column=16).value.capitalize()+'_'+ws.cell(row=cell.row, column=19).value.capitalize()+'(2).png', format="png")

		front = Image.open("./PNG_badge/02 (1).png")

		draw = ImageDraw.Draw(front)
		# font = ImageFont.truetype(<font-file>, <font-size>)
		font2 = ImageFont.truetype("./assets/font/vazir-font-v6.3.4/Vazir-Bold.ttf", 46, encoding="unic")
		font1 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", 86, encoding="unic")
		# font = ImageFont.truetype("./Lalezar-Regular.ttf", 26, encoding="unic")
		text1 = ws.cell(row=cell.row, column=16).value.capitalize()+' '+ws.cell(row=cell.row, column=19).value.capitalize()
		text2 = ws.cell(row=cell.row, column=20).value
		# unicode_text = unicode(text, "utf-8")
		# text = get_display(text)
		reshaped_text = arabic_reshaper.reshape(text1)
		final_text = get_display(reshaped_text)
		print final_text
		# draw.text((x, y),"Sample Text",(r,g,b))
		w1, h1 = draw.textsize(final_text, font=font1)
		font_size1 = 86
		if w1 > 900:
			w1, h1 = draw.textsize(final_text, font=font1) #get size again
			howbig1 = float((int(w1)-900))/900
			font_size1 -= (86*howbig1)
			font1 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", int(font_size1), encoding="unic")
			w1, h1 = draw.textsize(final_text, font=font1) #for drawing it with new size
			while not 900 >= w1 >= 850:
				font_size1 += font_size1*.01
				font1 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", int(font_size1), encoding="unic")
				w1, h1 = draw.textsize(final_text, font=font1)
		font1 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", int(font_size1), encoding="unic")
		w1, h1 = draw.textsize(final_text, font=font1) #for drawing it with new size
		draw.text(((945-w1)/2, 600),final_text,(255,255,255),font=font1)
		reshaped_text = arabic_reshaper.reshape(text2)
		final_text = get_display(reshaped_text)
		w2, h2 = draw.textsize(final_text, font=font2)

		font_size2 = 48
		if w2 > 460:
			w2, h2 = draw.textsize(final_text, font=font2) #get size again
			howbig2 = float((int(w2)-460))/460
			font_size2 -= (48*howbig2)
			font2 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", int(font_size2), encoding="unic")
			w2, h2 = draw.textsize(final_text, font=font2) #for drawing it with new size
			print ("** THE SIZE OF SECOND TEXT IS ", w2, " **")
			while not 460 >= w2 >= 350:
				font_size2 += font_size2*.01
				font2 = ImageFont.truetype("./assets/font/lato/Lato-Regular.ttf", int(font_size2), encoding="unic")
				w2, h2 = draw.textsize(final_text, font=font2)
		font2 = ImageFont.truetype("./assets/font/vazir-font-v6.3.4/Vazir-Bold.ttf", int(font_size2), encoding="unic")
		w2, h2 = draw.textsize(final_text, font=font2)
		draw.text(((945-w2)/2, 830),final_text,(255,255,255),font=font2)
		try:
			r = requests.get(ws.cell(row=cell.row, column=21).value, timeout=3)
			r.raise_for_status()
		except requests.exceptions.HTTPError as errh:
			print ("Http Error:",errh)
			continue
		except requests.exceptions.ConnectionError as errc:
			print ("Error Connecting:",errc)
			continue
		except requests.exceptions.Timeout as errt:
			print ("Timeout Error:",errt)
			continue
		except requests.exceptions.RequestException as err:
			print ("OOps: Something Else",err)
			continue

		if r.status_code == 200:
			headshot = Image.open(StringIO(r.content)).convert("RGBA")
			headshot = headshot.resize((353, 353));
			bigsize = (headshot.size[0] * 3, headshot.size[1] * 3)
			mask = Image.new('L', bigsize, 0)
			draw = ImageDraw.Draw(mask) 
			draw.ellipse((0, 0) + bigsize, fill=255)
			mask = mask.resize(headshot.size, Image.ANTIALIAS)
			headshot.putalpha(mask)

			output = ImageOps.fit(headshot, mask.size, centering=(0.5, 0.5))
			output.putalpha(mask)

			front.paste(headshot, (297, 204), headshot)
			front.save('./export/'+ws.cell(row=cell.row, column=16).value.capitalize()+'_'+ws.cell(row=cell.row, column=19).value.capitalize()+'(1).png', format="png")

